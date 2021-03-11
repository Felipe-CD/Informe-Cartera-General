import pandas as pd
from numpy import nan
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

sap = pd.read_excel("cartera 15 02 2021 virgen 1.XLS.xlsx",
                    skiprows=[0,1,2,3,4,5,6,7,9], usecols="B:AV")
cupos = pd.read_excel("cupos prepago.xlsx", skiprows=[0,1,2,4], usecols=[1,3,4,5,6,7,8,9])
limites = pd.ExcelFile("Copia de LIMITES  22012021.xlsx")
print(limites.sheet_names)
limites = pd.read_excel("Copia de LIMITES  22012021.xlsx", sheet_name="LIMITES")
lista = [x for x in sap.columns.to_list() if 'Cartera A 000' in x]
for i in lista:
    sap.drop(i, axis=1, inplace=True)

sap["Descripción cabecera pedido"] = sap["Descripción cabecera pedido"].str.lower()
sap["Descripción cabecera pedido"] = sap["Descripción cabecera pedido"].fillna("NA")

filtros = {
    "data": 
        {
            "name1": "Descripción cabecera pedido",
            "filt1": "castigo,valor presente neto,recarga en linea,vpn",
            "name2": "No. Referencia",
            "filt2": "CASTIGO"
        }
}
filtros["data"]["filt1"] = list(filtros["data"]["filt1"].split(","))
filtros["data"]["filt2"] = list(filtros["data"]["filt2"].split(","))
sap2 = []
for i in filtros["data"]["filt1"]:
    sap2.append(sap[sap["Descripción cabecera pedido"].str.contains(i)])
    sap.drop(sap[sap["Descripción cabecera pedido"].str.contains(i)].index, inplace=True)
for i in filtros["data"]["filt2"]:
    sap2.append(sap[(sap["No. Referencia"].notnull()) & (sap["No. Referencia"].str.contains(i))])
    sap.drop(sap[(sap["No. Referencia"].notnull()) & (sap["No. Referencia"].str.contains(i))].index, inplace=True)
sap.reset_index(drop=True, inplace=True)
#Las partidas quitadas con los filtros se colocan en un dataframe diferente para dejarlas en una hoja aparte
sap2 = pd.concat([sap2[i] for i in range(len(sap2))], axis=0)
sap2.reset_index(drop=True, inplace=True)
#Buscar acuerdos en el archivo limites para cada partida
sap["Status"] = sap["No. Identificación Fiscal"].map(limites.drop_duplicates("Nit").set_index("Nit")["ACUERDO"])
#Colocar "D" en cada partida que tiene la palabra "Cuota" en la columna descripcion cabecera pedido
sap.loc[sap["Descripción cabecera pedido"].str.contains("cuota") & sap["Ind. Cta Esp."].isnull(), ("Ind. Cta Esp.")] = "D"
#identificar cada agente que tenga por lo menos una D, y luego a todas las partidas de esos agentes colocarle “ACUERDO” en status
agentes_acuerdo = sap.loc[sap["Ind. Cta Esp."] == "D",("No. Identificación Fiscal")].drop_duplicates().to_list()
for i in agentes_acuerdo:
    sap.loc[sap["No. Identificación Fiscal"] == i,("Status")] = "ACUERDO"
#Reemolazar los vacios en "Status" por "Abierto"
sap.loc[sap["Status"].isnull(), ("Status")] = "ABIERTO"
#Si hay una "D" en la columna "Ind. Cta Esp." entonces se coloca un 20 en la columna "Producto"
#y se coloca "Acuerdo" en la columna "Tipo_producto"
sap.loc[sap["Ind. Cta Esp."] == "D", ("Producto")] = 20
sap.loc[sap["Ind. Cta Esp."] == "D", ("Tipo_Producto")] = "Acuerdo"
#Lo que no sea 20 o 10 en la columna producto entonces debe ser 18
sap.loc[(sap["Producto"] != 20.0) & (sap["Producto"] != 10.0), ("Producto")] = 18

#Filtro por dias de mora mayor a 1500 y se confirma que la suma de mayor a da 525.659.031 ¿Filtrar Latcom?
#incluir latcom y parar si la suma no da 525.659.031
#Colocar columna "Cartera total" en "Cartera no Vencido"
#Reemplazar por 0 las columnas "    MAyor a, Cartera Vencida" y por -1 la columna "Días Mora"
test_latcom = sap.loc[(sap["Descripción"].str.contains("LATCOM")) & (sap["Días Mora"] > 1500), ("         Mayor a")].sum()
if test_latcom != 525659031:
    flash('La suma de "   Mayor a" con el filtro (Días mora > 1500 y LATCOM) no es igual a $525.659.031')
else:
    flash('Filtro de "   MAyor a" verificado exitosamente')
sap.loc[(sap["Descripción"].str.contains("LATCOM")) & (sap["Días Mora"] > 1500), ("Cartera No Vencido")] = sap.loc[(sap["Descripción"].str.contains("LATCOM")) & (sap["Días Mora"] > 1500), ("     Cartera Total")]
sap.loc[(sap["Descripción"].str.contains("LATCOM")) & (sap["Días Mora"] > 1500), ("         Mayor a")] = 0
sap.loc[(sap["Descripción"].str.contains("LATCOM")) & (sap["Días Mora"] > 1500), (" Cartera Vencida")] = 0
sap.loc[(sap["Descripción"].str.contains("LATCOM")) & (sap["Días Mora"] > 1500), ("Días Mora")] = -1
#Buscar region en el archivo limites para cada partida y anexarla despues de columna 2 "No. Identificación Fiscal"
sap.insert(3, "Region", sap["No. Identificación Fiscal"].map(limites.drop_duplicates("Nit").set_index("Nit")["Nueva Region"]))
l1 = sap.loc[(sap["Status"] == "ACUERDO") & (sap["Ind. Cta Esp."].isnull()), ("Descripción")].drop_duplicates().to_list()
l2 = sap.loc[(sap["Status"] == "ACUERDO") & (sap["Ind. Cta Esp."] == "D"), ("Descripción")].drop_duplicates().to_list()
l3 = [x for x in l1 if x not in l2] #Distribuidores que se cambio de ACUERDO A ABIERTO
for i in l3:
    sap.loc[sap["Descripción"] == i, ("Status")] = "ABIERTO"

"""CREACION DE TABLAS DINAMICAS"""
#MEXICO
columns_to_analize1 = ['Cartera A 005 Días', 'Cartera A 010 Días', 'Cartera A 015 Días', 'Cartera A 020 Días', 
        'Cartera A 025 Días', 'Cartera A 030 Días']
columns_to_analize2 = ['Cartera No Vencido', 'Cartera A 060 Días', 'Cartera A 090 Días', 'Cartera A 120 Días', 
        '         Mayor a']
columns_to_analize11 = ['Cartera A 005 Días', '#5', 'Cartera A 010 Días', '#10', 'Cartera A 015 Días', '#15',
         'Cartera A 020 Días', '#20', 'Cartera A 025 Días', '#25', 'Cartera A 030 Días', '#30']
sap[columns_to_analize1] = sap[columns_to_analize1].replace({0:nan})
sap[columns_to_analize2] = sap[columns_to_analize2].replace({0:nan})
regiones = sap["Region"].drop_duplicates().to_list()
regiones.sort()
mexico = pd.DataFrame(columns=["Region"]+columns_to_analize11+columns_to_analize2)
cont = 0
for i in regiones:
    total1 = (sap.loc[sap["Region"] == i, (columns_to_analize1)].sum()/1000).round(0)
    cuenta1 = sap.loc[sap["Region"] == i, (['Descripción'] + columns_to_analize1)].groupby("Descripción").sum().replace({0:nan}).count()
    total2 = (sap.loc[sap["Region"] == i, (columns_to_analize2)].sum()/1000).round(0)
    row = [i]
    for j in range(len(total1)):
        row.append(total1[j])
        row.append(cuenta1[j])
    row = row + total2.to_list()
    mexico.loc[cont] = row
    cont += 1
mid = mexico["Cartera No Vencido"]
mexico.drop("Cartera No Vencido", axis=1, inplace=True)
mexico.insert(1, "ACTUAL", mid)
mexico.insert(14, "TOT 30", mexico["Cartera A 005 Días"] + mexico["Cartera A 010 Días"] + mexico["Cartera A 015 Días"] + mexico["Cartera A 020 Días"] + mexico["Cartera A 025 Días"] + mexico["Cartera A 030 Días"])

#CARTERA 120 DIAS
cartera_120 = sap.pivot_table(index="Descripción", values=["         Mayor a"], aggfunc=["sum"], columns=["Region"], margins=True)
cartera_120 = cartera_120[cartera_120[('sum', '         Mayor a', 'All')] > 0]

#Exportar sap, sap2, mexico y 120 dias a un archivo de excel en hojas diferentes
writer = pd.ExcelWriter("Informe_1.xlsx", engine="xlsxwriter")
workbook = writer.book
worksheet = workbook.add_worksheet("Base depurada")
writer.sheets["Base depurada"] = worksheet
worksheet.write_string(0, 0, f"Base depurada con los filtros especificados en el programa: {filtros['data']['filt1']}")
worksheet.write_string(1, 0, f"Los distribuidores a los cuales se les cambio el status de 'ACUERDO' a 'ABIERTO' son: {', '.join(str(x) for x in l3)}")
sap.to_excel(writer, sheet_name="Base depurada", startrow=2 , startcol=1, index=False)
worksheet = workbook.add_worksheet("Data extraida")
writer.sheets["Data extraida"] = worksheet
worksheet.write_string(0, 0, "Data extraida")
sap2.to_excel(writer, sheet_name="Data extraida", startrow=2 , startcol=1, index=False)
worksheet = workbook.add_worksheet("Informe Mexico")
writer.sheets["Informe Mexico"] = worksheet
worksheet.write_string(0, 1, "Comunicación Celular Comcel S.A")
worksheet.write_string(1, 1, "INTEGRACIÓN DE LA CUENTA POR COBRAR A")
worksheet.write_string(2, 1, "Canales de distribución regiones x-x")
worksheet.write_string(3, 1, "CIFRAS EN MILES DE PESOS")
mexico.to_excel(writer, sheet_name="Informe Mexico", startrow=10, startcol=1, index=False)
worksheet = workbook.add_worksheet("Cartera 120 dias")
writer.sheets["Cartera 120 dias"] = worksheet
worksheet.write_string(0, 0, "Informe cartera 120 dias")
cartera_120.to_excel(writer, sheet_name="Cartera 120 dias", startrow=2, startcol=1)
writer.save()

"""INFORME DE CARTERA GENERAL"""
def find_specific_row_cell(name,ws):
    """
    Función que retorna el numero de fila y columna de una celda qwe contiene el valor "name" en la hoja "ws"
    """
    for row in range(1, ws.max_row + 1):
        for column in "ABCDEFGHIJKL":
            cell_name = "{}{}".format(column, row)
            if ws[cell_name].value == name:
                return row, column

def limpiar_ajustar_rango(fila_inicio, fila_fin, columna_inicio, columna_fin, tamano_rango_viejo, tamano_rango_nuevo, ws):
    """
    Función que permite limpiar y ajustar un rango especificado de entrada segun el tamaño
    que tenga la data "dataframe" a exportar
    """
    rango_a_limpiar = "{}{}:{}{}".format(columna_inicio,fila_inicio,columna_fin,fila_fin)
    for row in ws[rango_a_limpiar]:
        for cell in row:
            cell.value = None
    if tamano_rango_viejo < tamano_rango_nuevo:
        ws.insert_rows(fila_inicio, (tamano_rango_nuevo - tamano_rango_viejo))
    elif tamano_rango_viejo > tamano_rango_nuevo:
        ws.delete_rows(fila_inicio, (tamano_rango_viejo - tamano_rango_nuevo))

#Agrupo y dejo solo columnas de dias de cartera 10, 20, 30
sap.insert(17, "Cartera a 10 Dias", sap[["Cartera A 005 Días","Cartera A 010 Días"]].sum(axis=1, min_count=1))
sap.insert(20, "Cartera a 20 Dias", sap[["Cartera A 015 Días","Cartera A 020 Días"]].sum(axis=1, min_count=1))
sap.insert(23, "Cartera a 30 Dias", sap[["Cartera A 025 Días","Cartera A 030 Días"]].sum(axis=1, min_count=1))
sap.drop(["Cartera A 005 Días","Cartera A 010 Días","Cartera A 015 Días","Cartera A 020 Días","Cartera A 025 Días","Cartera A 030 Días"], axis=1, inplace=True)
if sap["     Cartera Total"].sum() == sap["Cartera No Vencido"].sum() + sap["Cartera a 10 Dias"].sum() + sap["Cartera a 20 Dias"].sum() + sap["Cartera a 30 Dias"].sum() + sap["Cartera A 060 Días"].sum() + sap["Cartera A 090 Días"].sum() + sap["Cartera A 120 Días"].sum() + sap["         Mayor a"].sum():
    print("La suma desde 'Cartera total' hasta 'mayor a' da igual que 'Cartera Total'")
else:
    print("La suma desde 'Cartera total' hasta 'mayor a' NO da igual que 'Cartera Total'")

#CREACION DE TABLAS Y VARIABLES PARA MODIFICAR EL INFORME
#sap2 equivale a "Data_Conceptos_Excluyentes"
#par_exclu_inte equivale a la data en "Partidas Excluidas intereses"
par_exclu_inte = sap2.loc[
                        (sap2["Descripción cabecera pedido"].str.contains("valor presente neto")) | 
                        (sap2["Descripción cabecera pedido"].str.contains("vpn")),
                        ("No. de Cliente","Descripción","No. Identificación Fiscal","Cartera No Vencido","Días Mora"," Cartera Vencida","     Cartera Total")
]
par_exclu_inte["Descripción_2"] = "COMPENSACION INTERESES PRESTAMO DIS"
#suma_recargas equivale a la celda (Conceptos Recargas en Línea) de la hoja "Partidas Excluidas intereses"
suma_recargas = sap2.loc[
                        (sap2["Descripción cabecera pedido"].str.contains("recarga en linea")) |
                        (sap2["Descripción cabecera pedido"].str.contains("recarga $")) |
                        (sap2["Descripción cabecera pedido"].str.contains("recarga  $")),
                        ("     Cartera Total")
].sum()
#R1, R2, R3, R4, R5 para colocar ne las respectivas hojas del libro de excel
regiones = sap["Region"].drop_duplicates().to_list()
regiones.sort()
R1 = sap[sap["Region"] == regiones[0]]
R2 = sap[sap["Region"] == regiones[1]]
R3 = sap[sap["Region"] == regiones[2]]
R4 = sap[sap["Region"] == regiones[3]]
R5 = sap[sap["Region"] == regiones[4]]
#Data para las paginas "detalle otro concepto abierto" y "detalle kits"
otro_concepto_abierto = sap[sap["Producto"] == 18]
detalle_kits = sap[sap["Producto"] == 10]
#Tablas dinamicas
#Hoja Informe_Acuerdos
informe_acuerdos = sap[(sap["Status"] == "ACUERDO") & (sap["Ind. Cta Esp."].isnull())].pivot_table(
                    index=["No. de Cliente","No. Identificación Fiscal","Descripción"],
                    values=["Cartera No Vencido","Cartera a 10 Dias","Cartera a 20 Dias","Cartera a 30 Dias","Cartera A 060 Días","Cartera A 090 Días","Cartera A 120 Días","         Mayor a","     Cartera Total"], 
                    aggfunc=["sum"]) #Esta dataframe se debe imprimir con una fila antes, debido a que al quitar el header se imprime una despues
informe_acuerdos.columns = [j for i,j in informe_acuerdos.columns]
informe_acuerdos = informe_acuerdos[["Cartera No Vencido","Cartera a 10 Dias","Cartera a 20 Dias","Cartera a 30 Dias","Cartera A 060 Días","Cartera A 090 Días","Cartera A 120 Días","         Mayor a","     Cartera Total"]] #Reordenar el orden de las columnas
informe_acuerdos2 = sap[(sap["Status"] == "ACUERDO") & (sap["Ind. Cta Esp."] == "D")].pivot_table(
                    index=["No. de Cliente","No. Identificación Fiscal","Descripción"],
                    values=["     Cartera Total"], 
                    aggfunc=["sum"]) #Esta dataframe se debe imprimir con una fila antes, debido a que al quitar el header se imprime una despues
informe_acuerdos.insert(8, "Vencida", informe_acuerdos.iloc[:, 1:8].sum(axis=1))
informe_acuerdos["Acuerdo"] = informe_acuerdos2[("sum","     Cartera Total")]

#DEESDE ACA EMPECE A TRABAJAR SOLO EN EL ARCHIVO functions.py






#MODIFICACION DEL INFORME -------------------------------------------------------------
#Abrir el archivo de cartera general
book = load_workbook("xxx.xlsx")
#Eliminar todos los datos de una hoja
ws = book["Data_Conceptos_Excluyentes"]
ws.delete_rows(0, ws.max_row + 1)
#Preparacion para escribir en una(s) hojas
writer = pd.ExcelWriter('xxx.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#Eliminacion de celdas y rangos para que la data se ajuste

#Imprimir datframe en toda una hoja (Se puede especificar la fila y columna de inicio con startrow, startcol)
sap2.to_excel(writer, "Data_Conceptos_Excluyentes", index=False)
#guardar y cerrar el archivo
writer.save()


#CODIGO PARA ELIMINAR FILAS DE UNA HOJA ESPECIFICADA -------------------NOTAAAAAAAAAAAAAAAS
first=pd.DataFrame({'A':[1,1,1,1],
             'B':[2,2,2,2]})
book = load_workbook("test.xlsx")
writer = pd.ExcelWriter("test.xlsx", engine="openpyxl")
writer.book = book
writer.sheets = dict((wss.title, wss) for wss in book.worksheets)
ws = book["hoja de prueba"]
fila, columna = find_specific_row_cell("total",ws)
limpiar_ajustar_rango(2,(fila-1),"A","G",(fila-2),4,ws)
#Imprime el dataframe first en la columna d y la fila 8 del excel
first.to_excel(writer, "hoja de prueba", startrow=1, startcol=0, index=False, header=False)
writer.save()
#---------------------------------------------------------------------------------------------


#modificacion de la hoja partidas expluidas intereses, solo data de valor presente neto, falta colocar el valor de suma de recargas
book = load_workbook("xxx.xlsx")
ws = book["Partidas Excluidas intereses"]
fila, columna = find_specific_row_cell("Total",ws)
tamano_rango_viejo = fila - 7
rango_a_limpiar = "C7:J{}".format(fila)
for row in ws[rango_a_limpiar]:
    for cell in row:
        cell.value = None
if tamano_rango_viejo < len(par_exclu_inte):
    ws.insert_rows(7, (len(par_exclu_inte) - tamano_rango_viejo))
elif tamano_rango_viejo > len(par_exclu_inte):
    ws.delete_rows(7, (tamano_rango_viejo - len(par_exclu_inte)))

writer = pd.ExcelWriter('xxx.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
par_exclu_inte.to_excel(writer, "Partidas Excluidas intereses", index=False, header=False, startrow=6, startcol=2)
writer.save()